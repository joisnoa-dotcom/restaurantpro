import openpyxl
from openpyxl.styles import Font, Alignment
from app.models.setting import Setting
import io

def generate_sales_excel(payments):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Consultamos el nombre del restaurante dinámicamente
    setting = Setting.query.first()
    rest_name = setting.name if setting else "Ventas"
    # Excel tiene un límite de 31 caracteres para el nombre de la hoja
    ws.title = f"Reporte - {rest_name}"[:31] 

    headers = ['ID Pago', 'Fecha y Hora', 'Modalidad', 'Cliente', 'Método', 'Estado', 'Monto Base (S/)', 'Envío (S/)', 'Total Final (S/)', 'Utilidad (S/)', 'Platos Vendidos']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = openpyxl.styles.PatternFill(start_color="1B254B", end_color="1B254B", fill_type="solid")

    total_base = 0
    total_delivery = 0
    total_final = 0
    total_profit = 0

    for p in payments:
        platos = []
        costo_total_orden = 0
        for item in p.order.items:
            if item.status != 'cancelled':
                platos.append(f"{item.quantity}x {item.product.name}")
                costo_total_orden += float(item.product.cost or 0) * item.quantity
        detalle_platos = ", ".join(platos) if platos else "Sin detalles"

        # Lógica de Modalidad
        o = p.order
        if o.order_type == 'delivery':
            modalidad = "DELIVERY"
            cliente = f"{o.customer_name or 'N/A'} ({o.customer_phone or ''})"
        elif o.order_type == 'takeaway':
            modalidad = "RECOJO / LLEVAR"
            cliente = o.customer_name or 'N/A'
        else:
            modalidad = f"Mesa {o.table_rel.number if o.table_rel else 'N/A'}"
            cliente = "Comensal Local"

        monto_final = float(p.amount)
        monto_envio = float(o.delivery_fee) if o.order_type == 'delivery' else 0.00
        monto_base = monto_final - monto_envio
        utilidad = monto_base - costo_total_orden

        ws.append([
            p.id, 
            p.created_at.strftime('%Y-%m-%d %H:%M'), 
            modalidad,
            cliente,
            p.payment_method.upper(), 
            p.status.upper(),
            monto_base,
            monto_envio,
            monto_final,
            utilidad,
            detalle_platos
        ])
        
        if p.status == 'completed':
            total_base += monto_base
            total_delivery += monto_envio
            total_final += monto_final
            total_profit += utilidad

    ws.append(['', '', '', '', '', 'TOTALES:', total_base, total_delivery, total_final, total_profit, ''])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=9).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=10).font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output